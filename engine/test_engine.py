import math
import os
import pdb
import sys
import time
from datetime import datetime
from time import sleep
import numpy as np
import openpyxl
from matplotlib import pyplot
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from engine import convert_to_pdf
from lib import constant
from lib.utils import Key_Word_World

fuel_path = "fuel_n_kicker_Admin.xlsx"
sys_path = sys.path
cur_proj_path = sys_path[0]
fuel = os.path.join(cur_proj_path, fuel_path)

dt_strf = "%d_%m_%Y_%H_%M_%S"


def test_engine(driver):

    global start_time
    begin=time.time()
    cnt_ck = 0
    run_key = Key_Word_World(driver)

    def key_def(keyword, by=None, ele=None, val=None):
        if keyword == "open browser":
            val = str(val).lower()
            if val == "firefox" or val == "fire fox":
                dr_path = os.path.join(cur_proj_path, "driver\\geckodriver.exe")
            elif val == "ie":
                dr_path = os.path.join(cur_proj_path, "driver\\IEDriverServer.exe")
            elif val == "edge":
                dr_path = os.path.join(cur_proj_path, "driver\\msedgedriver.exe")
            else:
                dr_path = os.path.join(cur_proj_path, "driver\\chromedriver.exe")
            run_key.open_browser(dr_path, val)
        elif keyword == "load url":
            run_key.load_url(val)
        elif keyword == "enter text":
            run_key.enter_text(by, ele, val)
        elif keyword == "click on element":
            run_key.click_on_element(by, ele)
        elif keyword == "double click on element":
            run_key.double_click_on_element(by, ele)
        elif keyword == "choose value from element list":
            run_key.choose_value_from_element_list(by, ele, val)
        elif keyword == "verify value true":
            # run_key.move_to_the_element(by, ele)
            run_key.verify_value_true(by, ele, val)
        elif keyword == "verify value false":
            run_key.verify_value_false(by, ele, val)
        elif keyword == "press enter":
            run_key.press_enter_key()
        elif keyword == "press up key":
            run_key.press_arrow_up_key()
        elif keyword == "press down key":
            run_key.press_arrow_down_key()
        elif keyword == "press right":
            run_key.press_arrow_right_key()
        elif keyword == "press tab":
            run_key.press_tab_key()
        elif keyword == "press esc":
            run_key.press_escape_key()
        elif keyword == "kill browser":
            run_key.kill_browser()
        elif keyword == "sleep till":
            run_key.define_sleep(val)
        elif keyword == "zoom out":
            run_key.zoom_out()
        elif keyword == "clear text":
            run_key.clear_the_element(by, ele)
        elif keyword == "javascript click":
            run_key.javascript_click(by, ele)
        elif keyword == "element present false":
            run_key.check_element_not_present(by, ele)
        elif keyword == "element present true":
            run_key.check_element_present(by, ele)
        elif keyword == "element disabled True":
            run_key.check_element_disabled(by, ele)
        elif keyword == "element enabled True":
            run_key.check_element_enabled(by, ele)
        elif keyword == "click file three dots":
            run_key.click_three_dots(val)
        elif keyword == "click project three dots":
            run_key.click_project_three_dots(val)
        elif keyword == "click on project":
            run_key.click_on_project(val)
        elif keyword == "click TAN number":
            run_key.click_tan_number(val)
        elif keyword == "click file check box":
            run_key.click_file_check_box(val)
        elif keyword == "select text from pdf":
            run_key.click_and_drag(by, ele)
        elif keyword == "upload folder in AWS":
            run_key.pyautogui_to_upload_a_file(val)
        elif keyword == "mouse hover":
            run_key.move_to_find(by, ele)
        elif keyword == "move to element":
            run_key.move_to_the_element(by, ele)
        elif keyword == "accept alert":
            run_key.accept_alert()
        elif keyword == "prepare sample file":
            run_key.prepare_sample_file(ele, val)
        elif keyword == "update metadata file":
            run_key.update_metadata_file(by, ele, val)
        elif keyword == "update metadata file":
            run_key.update_metadata_file(by, ele, val)
        elif keyword == "wait till loading ends":
            run_key.wait_till_loading_ends()
        elif keyword == "verify value in list":
            run_key.move_to_the_element(by, ele)
            run_key.verify_value_in_list(by, ele, val)
        elif keyword == "match value in list":
            run_key.move_to_the_element(by, ele)
            run_key.match_value_in_list(by, ele, val)
        elif keyword == "click by coordinates":
            run_key.click_by_coordinates(ele, val)
        elif keyword == "get TAN status":
            run_key.get_TAN_status(ele, val)
        elif keyword == "select pagination as":
            run_key.select_pagination(val)
        elif keyword == "upload attachment file":
            run_key.upload_file(by, ele, val)
        elif keyword == "select specialty as":
            run_key.select_specialty_as(val)
        elif keyword == "delete specialty":
            run_key.delete_specialty(val)
        elif keyword == "verify audit due date":
            run_key.verify_audit_due_date()
        elif keyword == "fetch TAN number":
            run_key.print_tan()
        elif keyword == "verify gen flag in audit form":
            run_key.verify_generification_flag()
        else:
            print("Keyword is not available, Please add the requested keyword and rerun")

    wb = load_workbook(fuel, data_only=True)
    ws = wb['Engine']
    tc_id = []
    tc_desc = []
    data = []
    last_data = []
    rn_case = []
    status = "Pass"
    stp_des = []

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=2):
        case_id = row[0].value
        case_desc = row[1].value
        step_desc = row[2].value
        key_word = row[3].value
        locator = row[4].value
        element = row[5].value
        test_data = row[6].value
        run_step = row[7].value
        run_case = row[8].value

        if len(tc_id) == 0:
            tc_id.append(case_id)
            tc_desc.append(case_desc)
            rn_case.append(run_case)
            print()
            print("Running First test case: " + case_id)
            # start_time = int(datetime.now().strftime("%H%M%S"))
            start_time = math.trunc(time.time())

        elif case_id is not None:
            if case_id != tc_id[-1]:
                # end_time = int(datetime.now().strftime("%H%M%S"))
                # timer = end_time - start_time
                end_time = math.trunc(time.time())
                timer = end_time-start_time
                execution_time = str(str(math.trunc(timer)) + " Sec")
                last_data.insert(4, execution_time)
                datas = last_data
                data.append(datas)
                print("Moving to next test case: " + case_id)
                tc_id.append(case_id)
                tc_desc.append(case_desc)
                rn_case.append(run_case)
                # start_time = int(datetime.now().strftime("%H%M%S"))
                start_time = math.trunc(time.time())
                status = "Pass"
                last_data = []
                cnt_ck = cnt_ck + 1
                if "open browser" in key_word:
                    cnt_ck = 0

        if rn_case[-1] == "yes" and status == "Pass":
            if run_step == "yes":
                if cnt_ck > 0 and data != [] and case_id is not None:
                    s = -1
                    for x in data:
                        if "SKIP" in data[s]:
                            s = s - 1
                        else:
                            res = data[s]
                            break
                    if len(data) != (-(int(s + 1))) and "SKIP" not in res:

                        if "Failed to establish a new connection" in str(data[s]):
                            status = "Fail"
                            last_data = [tc_id[-1], tc_desc[-1], " - ", "Fail",
                                         " Continuation Scenario - Previous TC Failed ", " - "]
                        elif "Continuation Scenario - Previous TC Failed" in str(data[s]) \
                                and "open browser" not in key_word:
                            status = "Fail"
                            last_data = [tc_id[-1], tc_desc[-1], " - ", "Fail",
                                         " Continuation Scenario - Previous TC Failed ", " - "]
                        else:
                            try:
                                key_def(key_word, locator, element, test_data)
                                error = snap_path = "-"
                            except Exception as err:
                                stp_des.append(step_desc)
                                dt = datetime.now().strftime(dt_strf)
                                error = sys.exc_info()
                                print(error)
                                try:
                                    snap_path = os.path.join(cur_proj_path, "screen_shot_fail\\" + tc_id[-1] + " " +
                                                             stp_des[-1] + dt + ".png")
                                    run_key.save_snap_shot(snap_path)
                                except:
                                    snap_path = ""
                                status = "Fail"
                                run_key.kill_browser()
                            if error == "-" and key_word != "prepare sample file" and key_word != "kill browser":
                                snap_path = os.path.join(cur_proj_path, "screen_shot_pass\\" + tc_id[-1] + " " + str(
                                    start_time) + ".png")
                                run_key.save_snap_shot(snap_path)
                            last_data = [tc_id[-1], tc_desc[-1], step_desc, status, error, snap_path]
                    else:
                        try:
                            key_def(key_word, locator, element, test_data)
                            error = snap_path = "-"
                        except Exception as err:
                            stp_des.append(step_desc)
                            dt = datetime.now().strftime(dt_strf)
                            error = sys.exc_info()
                            print(error)
                            try:
                                snap_path = os.path.join(cur_proj_path, "screen_shot_fail\\" + tc_id[-1] + " " +
                                                         stp_des[-1] + dt + ".png")
                                run_key.save_snap_shot(snap_path)
                            except:
                                snap_path = ""
                            status = "Fail"
                            run_key.kill_browser()
                        if error == "-" and key_word != "prepare sample file" and key_word != "kill browser":
                            snap_path = os.path.join(cur_proj_path, "screen_shot_pass\\" + tc_id[-1] + " " + str(
                                start_time) + ".png")
                            run_key.save_snap_shot(snap_path)
                        last_data = [tc_id[-1], tc_desc[-1], step_desc, status, error, snap_path]
                else:
                    try:
                        key_def(key_word, locator, element, test_data)
                        error = snap_path = "-"
                    except Exception as err:
                        stp_des.append(step_desc)
                        dt = datetime.now().strftime(dt_strf)
                        error = sys.exc_info()
                        print(error)
                        try:
                            snap_path = os.path.join(cur_proj_path, "screen_shot_fail\\" + tc_id[-1] + " " +
                                                     stp_des[-1] + dt + ".png")
                            run_key.save_snap_shot(snap_path)
                        except:
                            snap_path = ""
                        status = "Fail"
                        run_key.kill_browser()
                    if error == "-" and key_word != "prepare sample file" and key_word != "kill browser":
                        snap_path = os.path.join(cur_proj_path, "screen_shot_pass\\" + tc_id[-1] + " " + str(
                            start_time) + ".png")
                        run_key.save_snap_shot(snap_path)
                    last_data = [tc_id[-1], tc_desc[-1], step_desc, status, error, snap_path]
        elif rn_case[-1] == "no":
            last_data = [tc_id[-1], tc_desc[-1], " - ", "SKIP", " - ", " - "]
    # end_time = int(datetime.now().strftime("%H%M%S"))
    end_time = math.trunc(time.time())
    end = time.time()
    timer = end_time - start_time
    execution_time = str(str(timer) + " Sec")
    last_data.insert(4, execution_time)
    datas = last_data
    data.append(datas)
    print(data)
    total_test_case = status_pass = status_fail = status_skip = Execution_Time = 0
    for x in data:
        total_test_case = total_test_case + 1
        if x[3] == "Pass":
            status_pass = status_pass + 1
        elif x[3] == "Fail":
            status_fail = status_fail + 1
        elif x[3] == "SKIP":
            status_skip = status_skip + 1
        timee = str(x[4]).split(" Sec")
        tme = float(timee[0])
        Execution_Time = Execution_Time + tme
    if status_pass > 0:
        if status_fail > 0:
            if status_skip > 0:
                labels = ["Pass", "Fail", "Skip"]
                colors = ['green', 'red', 'yellow']
                status_count = [status_pass, status_fail, status_skip]
            else:
                labels = ["Pass", "Fail"]
                colors = ['green', 'red']
                status_count = [status_pass, status_fail]
        else:
            labels = ["Pass", "Skip", "Fail"]
            colors = ['green']
            status_count = [status_pass, status_skip, status_fail]
    else:
        if status_skip > 0:
            labels = ["Fail", "Skip"]
            colors = ['red', 'yellow']
            status_count = [status_fail, status_skip]
        else:
            labels = ["Fail"]
            colors = ['red']
            status_count = [status_fail]

    y = np.array(status_count)
    pyplot.pie(y, labels=labels, startangle=90, colors=colors, autopct='%1.1f%%', shadow=False)
    # pyplot.legend()
    pyplot.title("Execution Time: " + str(math.trunc(end-begin)) + " Sec")
    dt = datetime.now().strftime(dt_strf)
    chart_path = os.path.join(cur_proj_path, ("reports\\charts\\" + dt + ".png"))
    pyplot.savefig(chart_path, bbox_inches="tight")

    rpt_excel_path = os.path.join(cur_proj_path, ("reports\\xlReport\\Report_" + dt + ".xlsx"))
    rpt_wb = openpyxl.Workbook()
    rpt_ws = rpt_wb.active
    rpt_ws.title = "Test Results"

    rpt_ws['A1'] = "Test Report - " + str(ws['D1'].value) + " : " + str(ws["D2"].value)
    rpt_ws['A3'] = str("Status Summary [Total No. of Test Case Executed: " + str(total_test_case) + " ]")
    rpt_ws['A12'] = "Detailed Test Report"

    h1_font = Font(name='TimesNewRoman',
                   size=15,
                   bold=True,
                   italic=False,
                   vertAlign=None,
                   underline='none',
                   strike=False,
                   color='FF000000')

    h2_font = Font(name='TimesNewRoman',
                   size=12.5,
                   bold=True,
                   italic=False,
                   vertAlign=None,
                   underline='none',
                   strike=False,
                   color='FF000000')

    t_h_font = Font(name='TimesNewRoman',
                    size=11.5,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    font = Font(name='TimesNewRoman',
                size=9.5,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
    fill = PatternFill(fill_type="lightUp",
                       start_color='00FFFF99',
                       end_color='00FFFF99')
    border = Border(left=Side(border_style="thin",
                              color='FF000000'),
                    right=Side(border_style="thin",
                               color='FF000000'),
                    top=Side(border_style="thin",
                             color='FF000000'),
                    bottom=Side(border_style="thin",
                                color='FF000000'),
                    diagonal=Side(border_style=None,
                                  color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style="thin",
                                 color='FFFF0000'),
                    vertical=Side(border_style=None,
                                  color='FF000000'),
                    horizontal=Side(border_style=None,
                                    color='FF000000')
                    )
    alignment = Alignment(horizontal='left',
                          vertical='center',
                          text_rotation=0,
                          wrap_text=True,
                          shrink_to_fit=True,
                          indent=0)
    rpt_ws.merge_cells('A1:G1')
    a1 = rpt_ws['A1']
    a1.font = h1_font
    a1.fill = fill

    rpt_ws.merge_cells('A3:G3')
    a3 = rpt_ws['A3']
    a3.font = h2_font
    a3.fill = fill
    rpt_ws.merge_cells('A12:G12')
    a12 = rpt_ws['A12']
    a12.font = h2_font
    a12.fill = fill

    img = Image(chart_path)
    img.height = 125
    img.width = 150
    rpt_ws.add_image(img, 'B5')

    rpt_ws['D5'] = "Status"
    rpt_ws['E5'] = "Count"
    b5 = rpt_ws['D5']
    c5 = rpt_ws['E5']
    b5.font = t_h_font
    b5.border = border
    b5.alignment = alignment
    c5.font = t_h_font
    c5.border = border
    c5.alignment = alignment

    status_zip = dict(zip(labels, status_count))
    row = 5
    for x in status_zip:
        row = 1 + row
        status_cell = "D" + str(row)
        count_cell = "E" + str(row)
        rpt_ws[status_cell] = x
        rpt_ws[count_cell] = status_zip[x]
        stts_cel = rpt_ws[status_cell]
        cnt_cell = rpt_ws[count_cell]
        stts_cel.font = font
        stts_cel.border = border
        stts_cel.alignment = alignment
        cnt_cell.font = font
        cnt_cell.border = border
        cnt_cell.alignment = alignment
    rpt_wb.save(rpt_excel_path)
    columns = ['Test Case ID', 'Test Case Description', 'Last Step Executed',
               'Test Case Status', 'Execution Time', 'Error Message', 'Screen Shot']
    i = 64
    for x in columns:
        rpt_col = str(chr(i + 1))
        cell = rpt_col + "13"
        rpt_ws[cell] = x
        temp = rpt_ws[cell]
        temp.font = t_h_font
        temp.border = border
        temp.alignment = alignment
        i = i + 1
    rpt_wb.save(rpt_excel_path)
    rpt_ws.column_dimensions["A"].width = len(columns[0]) + 3
    rpt_ws.column_dimensions["B"].width = len(columns[1]) + 3
    rpt_ws.column_dimensions["C"].width = len(columns[2]) + 5
    rpt_ws.column_dimensions["D"].width = len(columns[3]) + 1
    rpt_ws.column_dimensions["E"].width = len(columns[4]) + 1
    rpt_ws.column_dimensions["F"].width = len(columns[5]) + 8
    rpt_ws.column_dimensions["G"].width = len(columns[6]) + 25
    rpt_wb.save(rpt_excel_path)

    j = 13
    for x in data:
        i = 64
        j = j + 1
        for y in x:
            rpt_col = str(chr(i + 1))
            cell = rpt_col + str(j)
            if ".png" in str(y):
                img = Image(y)
                img.height = 109
                img.width = 230
                rpt_ws.row_dimensions[j].height = 85
                rpt_ws.add_image(img, cell)
            else:
                rpt_ws[cell] = str(y)
            txt = rpt_ws[cell]
            txt.font = font
            txt.border = border
            txt.alignment = alignment
            i = i + 1

    rpt_wb.save(rpt_excel_path)
    try:
        pdf_Path = os.path.join(cur_proj_path, ("reports\\pdfReports\\Report-" + dt + ".pdf"))
        convert_to_pdf(rpt_excel_path, pdf_Path)
        sleep(10)
        os.startfile(pdf_Path)
    except:
        print("Not able to convert excel report to PDF")
    os.startfile(rpt_excel_path)
    constant.FINISHED = True
